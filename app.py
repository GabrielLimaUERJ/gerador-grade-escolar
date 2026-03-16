import streamlit as st
import pandas as pd
import random
import json
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

st.set_page_config(page_title="Gerador de Grade Escolar", layout="wide")

# -------------------------
# GARANTIR QUE JSON EXISTE
# -------------------------
if not os.path.exists("professores.json"):
    with open("professores.json", "w") as f:
        json.dump({}, f)

st.title("📚 Gerador de Grade Escolar")

# -------------------------
# CONFIGURAÇÕES DE TEMPOS E TURMAS
# -------------------------
num_tempos = st.number_input("Número de tempos por dia", min_value=1, max_value=10, value=6)
num_turmas = st.number_input("Número de turmas", min_value=1, max_value=10, value=3)

dias = ["Seg", "Ter", "Qua", "Qui", "Sex"]
tempos = [f"{i+1:02}" for i in range(num_tempos)]
horarios = [f"{d}{t}" for d in dias for t in tempos]

turmas = [f"Turma {i}" for i in range(1, num_turmas + 1)]
st.write("Turmas consideradas:", turmas)

# -------------------------
# MEMÓRIA
# -------------------------
if "professores" not in st.session_state:
    st.session_state.professores = {}
    try:
        with open("professores.json", "r") as f:
            dados = json.load(f)
            if isinstance(dados, dict):
                st.session_state.professores = dados
                for chave, info in st.session_state.professores.items():
                    if 'turmas' not in info:
                        info['turmas'] = turmas.copy()
    except:
        st.session_state.professores = {}

for h in horarios:
    if h not in st.session_state:
        st.session_state[h] = False

# -------------------------
# FUNÇÕES DE MARCAÇÃO
# -------------------------
def marcar_todos():
    for h in horarios:
        st.session_state[h] = True

def limpar_todos():
    for h in horarios:
        st.session_state[h] = False

def marcar_dia(dia):
    for t in tempos:
        st.session_state[f"{dia}{t}"] = True

def limpar_dia(dia):
    for t in tempos:
        st.session_state[f"{dia}{t}"] = False

def salvar_professores():
    with open("professores.json", "w") as f:
        json.dump(st.session_state.professores, f)
    st.success("Professores salvos")

def carregar_professores():
    try:
        with open("professores.json", "r") as f:
            dados = json.load(f)
        if isinstance(dados, dict):
            st.session_state.professores = dados
            for chave, info in st.session_state.professores.items():
                if 'turmas' not in info:
                    info['turmas'] = turmas.copy()
            st.success("Professores carregados")
        else:
            st.session_state.professores = {}
    except:
        st.session_state.professores = {}
        st.warning("Arquivo vazio ou inexistente")

# -------------------------
# ADICIONAR PROFESSOR
# -------------------------
st.header("Adicionar professor")
nome = st.text_input("Nome do professor")
disciplina = st.text_input("Disciplina")
dois_tempos_seguidos = st.checkbox("Leciona dois tempos seguidos?")
tempos_semana = st.number_input("Quantos tempos este professor vai lecionar por semana?", min_value=1, max_value=num_tempos*len(dias))

st.subheader("Selecione as turmas para esta disciplina")
turmas_selecionadas = []
cols_turmas = st.columns(len(turmas))
for i, t in enumerate(turmas):
    if cols_turmas[i].checkbox(t):
        turmas_selecionadas.append(t)

col1, col2 = st.columns(2)
col1.button("Marcar todos os horários", on_click=marcar_todos)
col2.button("Limpar todos os horários", on_click=limpar_todos)

st.subheader("Disponibilidade")
cols = st.columns(len(dias))
for i, dia in enumerate(dias):
    with cols[i]:
        st.markdown(f"**{dia}**")
        for tempo in tempos:
            chave = f"{dia}{tempo}"
            st.checkbox(f"Tempo {tempo}", key=chave)
        c1, c2 = st.columns(2)
        c1.button("Marcar todos", key=f"marcar_{dia}", on_click=marcar_dia, args=(dia,))
        c2.button("Limpar", key=f"limpar_{dia}", on_click=limpar_dia, args=(dia,))

if st.button("Adicionar professor"):
    if nome and disciplina and turmas_selecionadas:
        disponibilidade = [h for h in horarios if st.session_state[h]]
        chave_prof = f"{nome} - {disciplina}"
        st.session_state.professores[chave_prof] = {
            "professor": nome,
            "disciplina": disciplina,
            "disponibilidade": disponibilidade,
            "dois_tempos": dois_tempos_seguidos,
            "tempos_semana": tempos_semana,
            "turmas": turmas_selecionadas
        }
        salvar_professores()
        st.success(f"{chave_prof} adicionado")
        st.rerun()
    else:
        st.warning("Preencha nome, disciplina e selecione pelo menos uma turma.")

# -------------------------
# LISTA DE PROFESSORES
# -------------------------
st.subheader("Professores cadastrados")
for chave_prof, info in st.session_state.professores.items():
    col1, col2 = st.columns([4,1])
    with col1:
        horarios_legiveis = [f"{h[:3]} Tempo {h[3:]}" for h in info.get("disponibilidade",[])]
        st.write(f"{chave_prof} → {', '.join(horarios_legiveis)} | Dois tempos: {info.get('dois_tempos', False)} | Aulas/semana: {info.get('tempos_semana',0)} | Turmas: {', '.join(info.get('turmas', []))}")
    with col2:
        if st.button("Remover", key=f"remover_{chave_prof}"):
            del st.session_state.professores[chave_prof]
            salvar_professores()
            st.rerun()

# -------------------------
# GERAR GRADE
# -------------------------
if st.button("Gerar grade"):
    professores = st.session_state.professores
    melhor_grade = None
    melhor_pontuacao = -9999
    melhor_contador_aulas = {}
    melhor_dois_tempos_nao = {}

    for tentativa in range(1000):
        grade = {}
        prof_ocupado = {}
        contador_aulas = {prof:0 for prof in professores}
        dois_tempos_nao_atendidos = {prof:0 for prof in professores}
        contador_por_turma = {prof:{t:0 for t in turmas} for prof in professores}

        for dia in dias:
            for idx, tempo in enumerate(tempos):
                turmas_random = turmas.copy()
                random.shuffle(turmas_random)
                for turma in turmas_random:
                    h = f"{dia}{tempo}"
                    candidatos = []

                    for prof, info in professores.items():
                        if contador_aulas[prof] >= info.get("tempos_semana",0):
                            continue
                        if h not in info.get("disponibilidade", []):
                            continue
                        if (prof,h) in prof_ocupado:
                            continue
                        if turma not in info.get("turmas", []):
                            continue

                        # Dois tempos consecutivos
                        if info.get("dois_tempos", False):
                            if idx+1 >= len(tempos):
                                dois_tempos_nao_atendidos[prof] += 1
                                continue
                            prox_h = f"{dia}{tempos[idx+1]}"
                            if prox_h not in info.get("disponibilidade", []) or (prof, prox_h) in prof_ocupado:
                                dois_tempos_nao_atendidos[prof] += 1
                                continue

                        candidatos.append(prof)

                    if not candidatos:
                        continue

                    random.shuffle(candidatos)
                    candidatos.sort(key=lambda p: (contador_por_turma[p][turma], contador_aulas[p]))
                    escolhido = candidatos[0]

                    # Aloca professor
                    grade[(turma,h)] = escolhido
                    prof_ocupado[(escolhido,h)] = True
                    contador_aulas[escolhido] += 1
                    contador_por_turma[escolhido][turma] += 1

                    # Segundo tempo consecutivo
                    if professores[escolhido].get("dois_tempos", False):
                        prox_h = f"{dia}{tempos[idx+1]}"
                        grade[(turma, prox_h)] = escolhido
                        prof_ocupado[(escolhido, prox_h)] = True
                        contador_aulas[escolhido] += 1
                        contador_por_turma[escolhido][turma] += 1

        # -------------------------
        # Pontuação
        # -------------------------
        total_preenchido = len(grade)
        total_faltando = sum(max(0, professores[p].get("tempos_semana",0) - contador_aulas[p]) for p in professores)
        total_dois_tempos_nao = sum(dois_tempos_nao_atendidos.values())
        turmas_preenchidas = [sum(1 for h in horarios if (turma,h) in grade) for turma in turmas]
        uniformidade = min(turmas_preenchidas) / max(1, max(turmas_preenchidas))

        # Penalidade: 3 ou mais aulas consecutivas na mesma turma
        penalidade_consecutivos = 0
        for turma in turmas:
            for dia in dias:
                cont = 0
                ultimo_prof = None
                for tempo in tempos:
                    h = f"{dia}{tempo}"
                    prof = grade.get((turma,h), None)
                    if prof == ultimo_prof and prof is not None:
                        cont += 1
                    else:
                        cont = 1
                    ultimo_prof = prof
                    if cont >= 3:
                        penalidade_consecutivos += 2

        pontuacao = 2*total_preenchido - 5*total_faltando - 3*total_dois_tempos_nao + 2*uniformidade - penalidade_consecutivos

        if pontuacao > melhor_pontuacao:
            melhor_pontuacao = pontuacao
            melhor_grade = grade
            melhor_contador_aulas = contador_aulas.copy()
            melhor_dois_tempos_nao = dois_tempos_nao_atendidos.copy()

    if melhor_grade is None:
        st.warning("Não foi possível gerar uma grade completa com os professores disponíveis.")
    else:
        grade = melhor_grade
        for turma in turmas:
            for h in horarios:
                if (turma,h) not in grade:
                    grade[(turma,h)] = ""

        # -------------------------
        # Mostrar tabela
        # -------------------------
        tabelas = {}
        for turma in turmas:
            tabela = []
            for tempo in tempos:
                linha = []
                for dia in dias:
                    chave = f"{dia}{tempo}"
                    linha.append(grade.get((turma,chave), ""))
                tabela.append(linha)
            df = pd.DataFrame(tabela, columns=dias, index=[f"Tempo {t}" for t in tempos])
            tabelas[turma] = df
            st.subheader(turma)
            st.table(df)

        # -------------------------
        # Exportar Excel
        # -------------------------
        arquivo = "grade_horarios.xlsx"
        with pd.ExcelWriter(arquivo, engine='openpyxl') as writer:
            for turma, df in tabelas.items():
                df.to_excel(writer, sheet_name=turma)

        wb = load_workbook(arquivo)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            for cell in ws[1]:
                cell.font = Font(bold=True)
        wb.save(arquivo)

        with open(arquivo, "rb") as f:
            st.download_button("📥 Baixar Excel", f, file_name="grade_horarios.xlsx")

    # -------------------------
    # Professores impossíveis
    # -------------------------
    impossiveis = {prof:max(0,professores[prof].get("tempos_semana",0)-melhor_contador_aulas.get(prof,0))
                   for prof in professores if melhor_contador_aulas.get(prof,0) < professores[prof].get("tempos_semana",0)}
    if impossiveis:
        st.error("Os seguintes professores não puderam ser totalmente encaixados na grade:")
        for prof,faltando in impossiveis.items():
            st.write(f"{prof} → faltam {faltando} tempos")
